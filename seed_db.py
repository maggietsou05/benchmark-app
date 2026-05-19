"""
Seed the SQLite benchmark.db from the reference Excel file in ../example/.

Run once after install (or any time you want to rebuild from the Excel).
This script is idempotent — it drops and recreates the DB on each run.
"""

import sqlite3
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
EXCEL_PATH = HERE.parent / "example" / "Benchmark leaflets pilot ClaudeAI.xlsx"
DB_PATH = HERE / "benchmark.db"

GROUP_NAMES = {
    "Picture/Display",
    "Connectivity",
    "Power Delivery",
    "Convinience",
    "Convenience",
    "Stand",
    "Power",
    "Dimension",
    "Weight",
    "Operating conditions",
    "Sustainability",
    "Compliance and standards",
    "Cabinet",
    "What's in the box?",
    "Not on Philips Leaflet",
}

GROUP_NORMALIZE = {"Convinience": "Convenience"}

TERMINATORS = {
    "Summary",
    "Leaflet Improvement Actions",
    "Product Development Summary",
    "Ports & Slots Reference",
    "Executive Summary",
}


def cell_to_str(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, timedelta):
        return f"[bad excel cell: {v}]"
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def parse_brand_model(full_name):
    parts = full_name.split(" ", 1)
    if len(parts) == 2:
        return parts[0], parts[1]
    return full_name, full_name


def parse_comparison_sheet(ws):
    product_a_full = None
    product_b_full = None
    current_group = None
    sort_order = 0
    specs_a = []
    specs_b = []
    in_comparison = False

    for row in ws.iter_rows(values_only=True):
        cells = list(row) + [None] * (8 - len(row))
        a, b, c, d, *_ = cells

        a_str = a.strip() if isinstance(a, str) else None
        b_str = b.strip() if isinstance(b, str) else None

        if not in_comparison:
            if b_str == "Picture/Display" and c and d:
                product_a_full = cell_to_str(c)
                product_b_full = cell_to_str(d)
                current_group = "Picture/Display"
                in_comparison = True
            continue

        # Only check terminators after we've entered comparison mode —
        # otherwise pre-table sections like "Executive Summary" bail us out early.
        if a_str in TERMINATORS:
            break

        if b_str in GROUP_NAMES and (c is None or cell_to_str(c) in (None, "")):
            current_group = GROUP_NORMALIZE.get(b_str, b_str)
            continue

        if not b_str:
            continue
        if b_str in GROUP_NAMES:
            continue

        sort_order += 1
        feature = b_str
        value_a = cell_to_str(c)
        value_b = cell_to_str(d)
        specs_a.append((current_group, feature, value_a, sort_order))
        specs_b.append((current_group, feature, value_b, sort_order))

    return product_a_full, product_b_full, specs_a, specs_b


# Lenovo P27Q-40 — placeholder data so dropdowns have a third option.
# Sparse on purpose; values not in this list will simply not appear in the DB.
LENOVO_P27Q40_SPECS = [
    ("Picture/Display", "LCD panel type:", "IPS"),
    ("Picture/Display", "Panel Size:", "27 inch / 68.6 cm"),
    ("Picture/Display", "Refresh Rate:", "60 Hz"),
    ("Picture/Display", "Maximum resolution:", "2560 x 1440 @ 60 Hz"),
    ("Picture/Display", "Pixel Density:", "109 PPI"),
    ("Picture/Display", "Brightness:", "350 cd/m²"),
    ("Picture/Display", "Contrast ratio (typical):", "1000:1"),
    ("Picture/Display", "Response time (typical):", "4 ms (extreme mode)"),
    ("Picture/Display", "Viewing angle: ", "178° (H) / 178° (V)"),
    ("Picture/Display", "Display colours: ", "16.7M (8 bits)"),
    ("Picture/Display", "Colour gamut (typical): ", "sRGB 99%"),
    ("Connectivity", "HDMI", "HDMI 1.4 x 1"),
    ("Connectivity", "DisplayPort", "DisplayPort 1.2 x 1"),
    ("Connectivity", "USB-C", "USB-C x 1 (DP Alt Mode, 75W PD)"),
    ("Connectivity", "USB Speed:", "USB 3.2 Gen 1 (5 Gbps)"),
    ("Connectivity", "USB Downstream", "USB-A x 4"),
    ("Connectivity", "Audio:", "Audio out 3.5 mm"),
    ("Connectivity", "RJ45:", "Gigabit Ethernet"),
    ("Power Delivery", "PD Version:", "USB PD 3.0"),
    ("Power Delivery", "Max power delivery: ", "75W"),
    ("Convenience", "Built-in Speakers:", "N/A"),
    ("Convenience", "MultiView:", "N/A"),
    ("Stand", "Height adjustment:", "135 mm"),
    ("Stand", "Pivot:", "-/+90 degree"),
    ("Stand", "Swivel:", "-45/+45 degree"),
    ("Stand", "Tilt:", "-5/+30 degree"),
    ("Power", "On mode:", "23 W (typ.)"),
    ("Power", "Standby mode:", "0.5 W"),
    ("Power", "Energy Label Class:", "F"),
    ("Compliance and standards", "Warranty:", "3 years"),
]


def main():
    if not EXCEL_PATH.exists():
        raise SystemExit(f"Excel file not found at {EXCEL_PATH}")

    print(f"Reading {EXCEL_PATH.name}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheet_name = next((s for s in wb.sheetnames if s != "Template"), None)
    if not sheet_name:
        raise SystemExit("No comparison sheet found (only the Template).")
    ws = wb[sheet_name]

    product_a, product_b, specs_a, specs_b = parse_comparison_sheet(ws)
    if not product_a or not product_b:
        raise SystemExit("Could not detect product names in comparison sheet.")

    print(f"  Found: {product_a} ({len(specs_a)} specs)")
    print(f"         {product_b} ({len(specs_b)} specs)")

    if DB_PATH.exists():
        DB_PATH.unlink()

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")
    conn.executescript(
        """
        CREATE TABLE products (
            id INTEGER PRIMARY KEY,
            brand TEXT NOT NULL,
            model TEXT NOT NULL UNIQUE,
            full_name TEXT NOT NULL,
            source_filename TEXT,
            source_pdf BLOB,
            website_url TEXT,
            ingested_by TEXT,
            ingested_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE specs (
            id INTEGER PRIMARY KEY,
            product_id INTEGER NOT NULL,
            group_name TEXT NOT NULL,
            feature TEXT NOT NULL,
            value TEXT,
            sort_order INTEGER,
            FOREIGN KEY (product_id) REFERENCES products(id) ON DELETE CASCADE
        );
        CREATE INDEX idx_specs_product ON specs(product_id);
        CREATE INDEX idx_specs_group ON specs(group_name);
        """
    )
    cur = conn.cursor()

    def insert_product(full_name, source_filename=None, ingested_by=None):
        brand, model = parse_brand_model(full_name)
        cur.execute(
            "INSERT INTO products (brand, model, full_name, source_filename, ingested_by) "
            "VALUES (?, ?, ?, ?, ?)",
            (brand, model, full_name, source_filename, ingested_by),
        )
        return cur.lastrowid

    def insert_specs(product_id, specs):
        for group, feature, value, sort_order in specs:
            if value is None or value == "":
                continue
            cur.execute(
                "INSERT INTO specs (product_id, group_name, feature, value, sort_order) "
                "VALUES (?, ?, ?, ?, ?)",
                (product_id, group, feature, value, sort_order),
            )

    excel_source = EXCEL_PATH.name
    pid_a = insert_product(product_a, source_filename=excel_source, ingested_by="Kevin Yang")
    pid_b = insert_product(product_b, source_filename=excel_source, ingested_by="Kevin Yang")
    insert_specs(pid_a, specs_a)
    insert_specs(pid_b, specs_b)

    pid_lenovo = insert_product(
        "Lenovo ThinkVision P27Q-40",
        source_filename=None,
        ingested_by="(seed_db.py placeholder)",
    )
    lenovo_with_order = [
        (group, feature, value, i + 1)
        for i, (group, feature, value) in enumerate(LENOVO_P27Q40_SPECS)
    ]
    insert_specs(pid_lenovo, lenovo_with_order)

    conn.commit()

    print("\nDB summary:")
    rows = cur.execute(
        "SELECT p.full_name, COUNT(s.id) FROM products p "
        "LEFT JOIN specs s ON s.product_id = p.id GROUP BY p.id ORDER BY p.id"
    ).fetchall()
    for full_name, count in rows:
        print(f"  {full_name}: {count} specs")
    conn.close()
    print(f"\nWrote {DB_PATH}")


if __name__ == "__main__":
    main()
