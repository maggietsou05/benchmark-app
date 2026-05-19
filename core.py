"""
Core operations for the monitor benchmark — pure functions, no Streamlit.

Three primitives that drive every entry point in the app:

  extract_monitor_from_pdf  PDF bytes → MonitorSpec (Claude Haiku)
  search_monitors           query string → list[MonitorSummary]
  generate_comparison       two model ids + template → Excel bytes (Claude Sonnet)

Keeping these out of the UI layer means a CLI, scheduled job, Power BI prep
step, or alternate front-end can call them the same way Streamlit does.
"""

import base64
import json
import re
import sqlite3
from html.parser import HTMLParser
from io import BytesIO
from pathlib import Path
from typing import Literal

import anthropic
import httpx
import openpyxl
import pymupdf
import yaml
from dotenv import load_dotenv
from pydantic import BaseModel, Field

import xlsx_styles

load_dotenv()

HERE = Path(__file__).parent
DB_PATH = HERE / "benchmark.db"
PROMPTS_DIR = HERE / "prompts"

# Canonical group names — display order and the allowed enum values for
# Claude's extraction schema. Order matters: it controls how specs are
# rendered in the UI.
GROUP_NAMES: list[str] = [
    "Picture/Display",
    "Connectivity",
    "Power Delivery",
    "Convenience",
    "Stand",
    "Power",
    "Dimension",
    "Weight",
    "Operating conditions",
    "Sustainability",
    "Compliance and standards",
    "Cabinet",
    "What's in the box?",
    "Not on Philips Leaflet",
]

GROUP_LITERAL = Literal[
    "Picture/Display", "Connectivity", "Power Delivery", "Convenience",
    "Stand", "Power", "Dimension", "Weight", "Operating conditions",
    "Sustainability", "Compliance and standards", "Cabinet",
    "What's in the box?", "Not on Philips Leaflet",
]


# ─── Pydantic models ───────────────────────────────────────────────

class ExtractedSpec(BaseModel):
    group: GROUP_LITERAL
    feature: str
    value: str  # "Not listed" if absent from the leaflet


class VisualObservation(BaseModel):
    """A physical feature visible in a leaflet diagram but absent from the
    spec text — e.g. pop-out USB hub, joystick, KVM button."""
    feature: str           # "Pop-out downstream USB hub"
    visual_cue: str        # what was seen, with page reference
    confidence: Literal["high", "medium", "low"]
    suggested_group: GROUP_LITERAL


class MonitorSpec(BaseModel):
    """Structured monitor specs as extracted from a leaflet."""
    brand: str
    model: str
    specs: list[ExtractedSpec]
    visual_observations: list[VisualObservation] = Field(default_factory=list)


class MonitorSummary(BaseModel):
    """One row from the products table — what `search_monitors` returns."""
    id: int
    brand: str
    model: str
    full_name: str
    spec_count: int
    source_filename: str | None = None
    website_url: str | None = None
    ingested_by: str | None = None
    ingested_at: str | None = None
    updated_at: str | None = None
    has_pdf: bool = False


class ComparisonRow(BaseModel):
    group: str
    feature: str
    philips_value: str
    # One value per competitor, parallel to the competitor_labels passed to
    # the Excel writer. Order matches the order the user picked competitors
    # in the Generate tab.
    competitor_values: list[str]
    # "Philips" / "Tie" / "Investigate" / or the full_name of the winning
    # competitor. With N competitors, "Competitor" alone is ambiguous, so
    # the agent names the winner explicitly.
    verdict: str
    notes: str


class Comparison(BaseModel):
    rows: list[ComparisonRow]
    summary: str


# ─── Block-based comparison output ─────────────────────────────────
# Used by the three pre-built strategic templates (R&D, B2C marketing,
# B2B marketing). Each block has its own column schema declared in the
# YAML; the LLM emits content per block keyed by `block_key`.

class BlockSpec(BaseModel):
    """Block layout parsed from the YAML template."""
    key: str
    name: str
    columns: list[str]            # ordered column keys
    headers: list[str]            # parallel display headers


class BlockContent(BaseModel):
    """One block's content as emitted by the LLM."""
    block_key: str
    narrative: str
    rows: list[dict[str, str]] = Field(default_factory=list)


class BlockComparison(BaseModel):
    """A full block-based comparison output."""
    blocks: list[BlockContent]
    summary_narrative: str = ""


class TemplateInfo(BaseModel):
    """Structured view of a template YAML — drives UI and render decisions."""
    name: str
    description: str = ""
    groups: list[str] | None = None
    system_prompt: str
    output_blocks: list[BlockSpec] | None = None       # None = flat (legacy)
    include_summary_narrative: bool = False             # R&D Block 4
    show_url_fields: bool = False                       # marketing + custom
    download_mode: Literal["workbook", "per_block"] = "workbook"
    raw: dict = Field(default_factory=dict)


# ─── DB ────────────────────────────────────────────────────────────

def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def _flat_specs(product_id: int) -> dict[str, dict[str, str]]:
    """{group: {feature: value}} — the shape Claude wants for comparison input."""
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT group_name, feature, value FROM specs "
            "WHERE product_id = ? ORDER BY sort_order",
            (product_id,),
        ).fetchall()
    grouped: dict[str, dict[str, str]] = {}
    for group, feature, value in rows:
        grouped.setdefault(group, {})[feature] = value or ""
    return grouped


def get_monitor_specs(model: str) -> dict:
    """Full record for one monitor by model number.

    Returns a dict with brand/model/full_name and `specs` as
    {group: {feature: value}}. Raises ValueError if the model is unknown.
    """
    with get_conn() as conn:
        row = conn.execute(
            "SELECT id, brand, model, full_name FROM products WHERE model = ?",
            (model,),
        ).fetchone()
    if not row:
        raise ValueError(f"No monitor with model '{model}' in the database.")
    product_id, brand, model_, full_name = row
    return {
        "brand": brand,
        "model": model_,
        "full_name": full_name,
        "specs": _flat_specs(product_id),
    }


# ─── Templates ─────────────────────────────────────────────────────

def load_templates() -> dict:
    """Comparison templates keyed by filename stem."""
    return {
        f.stem: yaml.safe_load(f.read_text(encoding="utf-8"))
        for f in PROMPTS_DIR.glob("*.yaml")
    }


def resolve_template(template: str | dict) -> dict:
    """Look up a template by filename stem, or pass through a dict directly.

    Lets call sites accept either a saved template ("rd_deepdive") or an
    ad-hoc one built from the Custom View form. A dict must include
    `system_prompt`; everything else is optional.
    """
    if isinstance(template, dict):
        if "system_prompt" not in template:
            raise ValueError("Custom template dict must include 'system_prompt'.")
        return template
    templates = load_templates()
    if template not in templates:
        raise ValueError(
            f"Unknown template '{template}'. Available: {sorted(templates)}"
        )
    return templates[template]


# Skeleton wrapped around the user's three custom-view fields. Hard-codes
# the verdict scheme, summary length, and spec-group inclusion (all groups)
# so non-tech users only fill in what's actually decision-relevant.
_CUSTOM_VIEW_PROMPT = """\
You are writing a competitive monitor comparison for the Philips/MMD team.

AUDIENCE: {audience}

GOAL: {goal}
{other_details_section}
You are comparing ONE Philips/OBM monitor against ONE OR MORE competitor
monitors. For each meaningful spec, output one row. The row carries
Philips's value, every competitor's value (one entry per competitor in
the order given), and a verdict.

Verdict scheme:
- "Philips" — Philips beats every competitor on this spec
- the full_name of a specific competitor — that competitor beats every
  other monitor (including Philips) on this spec
- "Tie" — at least two monitors are equivalent and lead
- "Investigate" — subjective, marketing-only, or any side did not list
  the value (data gap requiring human review)

Notes should advance the GOAL above — frame each spec in terms of what
helps the audience decide, not just restate the value. Skip rows where
all sides are trivially identical or unspecified.

In the summary:
- 5-8 sentences total
- Address the audience defined above
- Lead with the strongest decision-driver for that audience, not the
  most impressive raw spec
- If multiple competitors are present, call out which competitor poses
  which kind of threat to Philips
"""


def build_custom_template(audience: str, goal: str, other_details: str = "") -> dict:
    """In-memory template built from the Custom View form fields.

    Audience and goal are required (caller validates). Other details are
    optional free text — if blank, the section is omitted from the prompt
    rather than left as an empty header.
    """
    other_details = (other_details or "").strip()
    other_details_section = (
        f"\nOTHER DETAILS: {other_details}\n" if other_details else ""
    )
    system_prompt = _CUSTOM_VIEW_PROMPT.format(
        audience=audience.strip(),
        goal=goal.strip(),
        other_details_section=other_details_section,
    )
    return {
        "name": "Custom View",
        "description": f"User-defined view for {audience.strip()}.",
        "system_prompt": system_prompt,
        # Custom view: URL fields optional (user can enrich the LLM context
        # with web copy if they want a marketing-flavoured custom view).
        "show_url_fields": True,
        # No "groups" key → all spec groups included.
        # No "output_blocks" → legacy flat comparison.
    }


def parse_template_info(template: str | dict) -> TemplateInfo:
    """Parse a template (filename stem or in-memory dict) into TemplateInfo.

    Block templates declare `output_blocks` in the YAML; without it the
    template is treated as a flat (legacy) comparison.
    """
    raw = resolve_template(template)
    output_blocks_raw = raw.get("output_blocks")
    output_blocks: list[BlockSpec] | None = None
    if output_blocks_raw:
        output_blocks = [
            BlockSpec(
                key=b["key"],
                name=b["name"],
                columns=list(b["columns"]),
                headers=list(b["headers"]),
            )
            for b in output_blocks_raw
        ]
    return TemplateInfo(
        name=raw.get("name", "View"),
        description=raw.get("description", ""),
        groups=raw.get("groups"),
        system_prompt=raw["system_prompt"],
        output_blocks=output_blocks,
        include_summary_narrative=bool(raw.get("include_summary_narrative", False)),
        show_url_fields=bool(raw.get("show_url_fields", False)),
        download_mode=raw.get("download_mode", "workbook"),
        raw=raw,
    )


# ─── URL fetching (for marketing / custom templates) ───────────────

class _TextExtractor(HTMLParser):
    """Strip script/style/head and concatenate visible text from HTML.

    Only includes elements that *contain text we want to skip* — never
    void elements (meta, link, br, hr, img...). Void elements have no
    closing tag in HTML5 so `handle_endtag` never fires for them, which
    would leave `_skip_depth` permanently above zero and silently swallow
    the entire <body>.
    """
    _SKIP = {"script", "style", "noscript", "head"}

    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []
        self._skip_depth = 0

    def handle_starttag(self, tag: str, attrs: list) -> None:
        if tag in self._SKIP:
            self._skip_depth += 1

    def handle_endtag(self, tag: str) -> None:
        if tag in self._SKIP and self._skip_depth:
            self._skip_depth -= 1

    def handle_data(self, data: str) -> None:
        if self._skip_depth == 0:
            text = data.strip()
            if text:
                self.parts.append(text)


def fetch_url_text(url: str, max_chars: int = 50_000) -> str:
    """Fetch a product-page URL and return its visible text.

    Used at generate-time by marketing/custom templates so the LLM can
    compare each side's customer-facing positioning (not just leaflet
    specs). Best-effort: raises ValueError on any HTTP or parse failure
    so the caller can decide whether to degrade gracefully.
    """
    if not url or not url.strip():
        raise ValueError("URL is empty.")
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (compatible; MonitorBenchmarkBot/1.0; internal-MMD)"
        )
    }
    try:
        resp = httpx.get(url, follow_redirects=True, timeout=15.0, headers=headers)
        resp.raise_for_status()
    except httpx.HTTPError as e:
        raise ValueError(f"Could not fetch {url}: {e}") from e
    parser = _TextExtractor()
    try:
        parser.feed(resp.text)
    except Exception as e:
        raise ValueError(f"Could not parse HTML from {url}: {e}") from e
    text = "\n".join(parser.parts)
    text = re.sub(r"\n{3,}", "\n\n", text).strip()
    return text[:max_chars]


# ─── Claude ────────────────────────────────────────────────────────

_client = anthropic.Anthropic()

EXTRACTION_PROMPT = """You are extracting structured data from a monitor product leaflet.

You receive both the EXTRACTED TEXT and the RENDERED PAGE IMAGES.
Use the text for spec values. Use the images to spot physical features
visible only in diagrams (often missing from the spec table).

Identify:
1. The monitor's BRAND (e.g. "Philips", "Dell", "Lenovo")
2. The monitor's MODEL number (e.g. "27B2U6903", "U2725QE")
3. SPECS — all technical specifications, organized into the 14 groups:
   Picture/Display, Connectivity, Power Delivery, Convenience, Stand, Power,
   Dimension, Weight, Operating conditions, Sustainability,
   Compliance and standards, Cabinet, What's in the box?, Not on Philips Leaflet
   (use "Not on Philips Leaflet" as catch-all for specs that don't fit elsewhere)
4. VISUAL OBSERVATIONS — physical features visible in the diagrams but
   NOT written anywhere in the spec text. Look specifically for:
   - Inset / close-up illustration of a port cluster → likely indicates a
     pop-out, pop-up, or hidden-hub mechanism on the bezel or underside
   - Numbered callouts on a product photo pointing to a flap, drawer,
     joystick, KVM button, or retractable element
   - Cable-management cutout on the stand
   - Cover plates or removable panels
   - Indicator-light placement
   - VESA mount detail visible on a rear-view diagram
   For each, describe what visual cue you saw and on which page.

CRITICAL RULES:
- Spec values: if not explicitly stated in the TEXT, set value to "Not listed".
- Do NOT infer or guess spec values from images. Images are ONLY used to
  populate visual_observations, never to fill in spec values.
- Use exact text/numbers/units from the leaflet — preserve formatting
  like "350 cd/m²", "60 Hz", "1000:1".
- Each spec gets ONE row in `specs`.
- If a feature is BOTH in the text AND visible in a diagram, put it ONLY
  in `specs`. visual_observations is exclusively for diagram-only features.
- Use industry-standard feature names ("Refresh Rate", "USB-C",
  "Brightness", "Contrast ratio (typical)").
- Confidence on a visual observation:
    high   = clear, unambiguous indicator (e.g. labeled "pop-out hub")
    medium = strong suggestion (e.g. inset close-up of a port cluster)
    low    = uncertain (could be just a "look here" zoom, not a hidden mechanism)
"""


def _extract_pdf_text(pdf_bytes: bytes) -> str:
    doc = pymupdf.open(stream=pdf_bytes, filetype="pdf")
    text = "\n\n".join(page.get_text() for page in doc)
    doc.close()
    return text


def _render_pdf_pages(pdf_bytes: bytes, dpi: int = 150) -> list[bytes]:
    """Render each PDF page to PNG bytes for Claude's vision input.

    150 DPI is a sweet spot — small port labels and numbered callouts stay
    readable, while keeping image-token cost in cents per leaflet.
    """
    doc = pymupdf.open(stream=pdf_bytes, filetype="pdf")
    images = [page.get_pixmap(dpi=dpi).tobytes("png") for page in doc]
    doc.close()
    return images


# ─── The three core operations ─────────────────────────────────────

def extract_monitor_from_pdf(pdf_bytes: bytes) -> MonitorSpec:
    """Parse a PDF leaflet's bytes into a structured MonitorSpec via Claude.

    Sends both extracted text AND rendered page images. Text drives spec
    values; images let Claude flag physical features visible only in
    diagrams (pop-out hubs, joysticks, cable management, etc.) as
    visual_observations for human review.

    Raises ValueError if the PDF has no extractable text (likely a scanned
    image — needs OCR first). Propagates anthropic.APIError on API failure.
    """
    pdf_text = _extract_pdf_text(pdf_bytes)
    if not pdf_text.strip():
        raise ValueError(
            "PDF contains no extractable text — likely scanned. "
            "Run OCR first, then re-upload."
        )
    page_images = _render_pdf_pages(pdf_bytes)

    content: list = [{"type": "text", "text": f"Leaflet text:\n\n{pdf_text}"}]
    for i, img in enumerate(page_images, 1):
        content.append({"type": "text", "text": f"\n--- Page {i} image ---"})
        content.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": "image/png",
                "data": base64.b64encode(img).decode(),
            },
        })

    response = _client.messages.parse(
        model="claude-sonnet-4-6",
        max_tokens=8000,
        system=[{
            "type": "text",
            "text": EXTRACTION_PROMPT,
            "cache_control": {"type": "ephemeral"},
        }],
        messages=[{"role": "user", "content": content}],
        output_format=MonitorSpec,
    )
    return response.parsed_output


def search_monitors(query: str) -> list[MonitorSummary]:
    """Monitors whose brand, model, or full_name contains `query`.

    Empty query returns everything. Match is case-insensitive substring.
    Ordered by brand, then model.
    """
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT p.id, p.brand, p.model, p.full_name,
                   COUNT(s.id) AS spec_count,
                   p.source_filename, p.website_url, p.ingested_by,
                   p.ingested_at, p.updated_at,
                   p.source_pdf IS NOT NULL AS has_pdf
            FROM products p
            LEFT JOIN specs s ON s.product_id = p.id
            GROUP BY p.id
            ORDER BY p.brand, p.model
            """
        ).fetchall()
    cols = ["id", "brand", "model", "full_name", "spec_count",
            "source_filename", "website_url", "ingested_by",
            "ingested_at", "updated_at", "has_pdf"]
    summaries = [MonitorSummary(**dict(zip(cols, r))) for r in rows]
    q = query.strip().lower()
    if not q:
        return summaries
    return [
        m for m in summaries
        if q in m.brand.lower()
        or q in m.model.lower()
        or q in m.full_name.lower()
    ]


def compute_comparison(
    model_a: str, competitor_models: list[str], template: str | dict
) -> tuple[Comparison, str, list[str]]:
    """Call Claude to compare one Philips/OBM monitor against N competitors.

    `template` is either a filename stem (e.g. "rd_deepdive") or an
    in-memory template dict (e.g. from build_custom_template). Returns
    (comparison, label_a, competitor_labels) — the structured result plus
    the display names. Format-agnostic; see comparison_to_excel for the
    .xlsx writer.

    Raises ValueError if any monitor is missing, the template is unknown,
    or the Philips model also appears in the competitor list. Propagates
    anthropic.APIError on failure.
    """
    if not competitor_models:
        raise ValueError("Pick at least one competitor.")
    if model_a in competitor_models:
        raise ValueError("Philips model also appears in the competitor list.")
    if len(set(competitor_models)) != len(competitor_models):
        raise ValueError("Pick distinct competitors — duplicates are not allowed.")

    tmpl = resolve_template(template)

    all_models = [model_a] + competitor_models
    with get_conn() as conn:
        placeholders = ",".join("?" * len(all_models))
        rows = conn.execute(
            f"SELECT id, model, full_name FROM products WHERE model IN ({placeholders})",
            all_models,
        ).fetchall()
    by_model = {model: (pid, full_name) for pid, model, full_name in rows}
    for m in all_models:
        if m not in by_model:
            raise ValueError(f"Monitor with model '{m}' not found in DB.")

    pid_a, a_label = by_model[model_a]
    specs_a = _flat_specs(pid_a)

    competitor_data: list[tuple[str, dict]] = []
    for cm in competitor_models:
        pid_c, c_label = by_model[cm]
        specs_c = _flat_specs(pid_c)
        if tmpl.get("groups"):
            specs_c = {g: specs_c.get(g, {}) for g in tmpl["groups"] if specs_c.get(g)}
        competitor_data.append((c_label, specs_c))

    if tmpl.get("groups"):
        specs_a = {g: specs_a.get(g, {}) for g in tmpl["groups"] if specs_a.get(g)}

    competitor_blocks = "\n\n".join(
        f"Competitor {i} ({c_label}):\n```json\n"
        f"{json.dumps(c_specs, indent=2, ensure_ascii=False)}\n```"
        for i, (c_label, c_specs) in enumerate(competitor_data, start=1)
    )

    user_msg = (
        f"Generate a competitive comparison. ONE Philips/OBM monitor against "
        f"{len(competitor_data)} competitor monitor(s).\n\n"
        f"Philips/OBM ({a_label}):\n```json\n"
        f"{json.dumps(specs_a, indent=2, ensure_ascii=False)}\n```\n\n"
        f"{competitor_blocks}\n\n"
        "For each meaningful feature, output one row. `competitor_values` "
        "must be a list with one entry per competitor, in the same order "
        "as the competitors above.\n\n"
        "Verdict scheme:\n"
        "- 'Philips' if Philips/OBM beats every competitor\n"
        "- the full_name of a competitor if it beats every other monitor\n"
        "  (including Philips)\n"
        "- 'Tie' if equivalent across the leaders\n"
        "- 'Investigate' if subjective, marketing-only (e.g. dynamic "
        "contrast), or any side did not list the value (data gap requiring "
        "human review)\n"
    )

    response = _client.messages.parse(
        model="claude-sonnet-4-6",
        max_tokens=16000,
        system=[{
            "type": "text",
            "text": tmpl["system_prompt"],
            "cache_control": {"type": "ephemeral"},
        }],
        messages=[{"role": "user", "content": user_msg}],
        output_format=Comparison,
    )
    competitor_labels = [label for label, _ in competitor_data]
    return response.parsed_output, a_label, competitor_labels


_SHEET_NAME_INVALID = re.compile(r'[\[\]\:\*\?\/\\]')


def _safe_sheet_name(name: str) -> str:
    """Sanitize a name to fit Excel's 31-char limit and forbidden-char rules."""
    cleaned = _SHEET_NAME_INVALID.sub("-", name).strip()
    return cleaned[:31] or "Sheet"


def _resolve_headers(
    headers: list[str], label_a: str, competitor_labels: list[str]
) -> list[str]:
    """Replace generic 'Philips' / 'Competitor' headers with actual full names.

    'Philips' (literal or `{philips}` placeholder) always swaps for label_a.
    'Competitor' / `{competitor}` swaps for the first competitor's label — a
    fallback for the rare YAML header that mentions the competitor outside
    a per-competitor column. The per-competitor columns themselves are
    expanded separately by `expand_columns_for_competitors`.
    """
    out = []
    primary = competitor_labels[0] if competitor_labels else ""
    for h in headers:
        h = h.replace("{philips}", label_a).replace("{competitor}", primary)
        if h == "Philips":
            h = label_a
        elif h == "Competitor":
            h = primary
        out.append(h)
    return out


# Column keys that should be expanded into one column per competitor at render
# time. The agent emits row-dict keys like "competitor_value_1", "_2", etc.
_PER_COMPETITOR_PREFIX = "competitor_"


def _is_per_competitor_key(key: str) -> bool:
    return key == "competitor" or key.startswith(_PER_COMPETITOR_PREFIX)


def expand_columns_for_competitors(
    column_keys: list[str],
    column_headers: list[str],
    competitor_labels: list[str],
) -> tuple[list[str], list[str]]:
    """Expand any 'competitor_*' column into one column per competitor.

    Example: ['feature', 'competitor_value', 'change_type'] with two
    competitors → ['feature', 'competitor_value_1', 'competitor_value_2',
    'change_type']. Each expanded column's header is the competitor's
    full_name. Non-competitor columns pass through unchanged.
    """
    out_keys: list[str] = []
    out_headers: list[str] = []
    for key, header in zip(column_keys, column_headers):
        if _is_per_competitor_key(key):
            for idx, c_label in enumerate(competitor_labels, start=1):
                out_keys.append(f"{key}_{idx}")
                out_headers.append(c_label)
        else:
            out_keys.append(key)
            out_headers.append(header)
    return out_keys, out_headers


def _comparison_row_to_dict(row: ComparisonRow) -> dict[str, str]:
    """Flatten a ComparisonRow into the per-competitor key shape the xlsx
    writer expects ('competitor_value_1', '_2', ...).
    """
    d: dict[str, str] = {
        "group": row.group,
        "feature": row.feature,
        "philips_value": row.philips_value,
        "verdict": row.verdict,
        "notes": row.notes,
    }
    for i, val in enumerate(row.competitor_values, start=1):
        d[f"competitor_value_{i}"] = "" if val is None else str(val)
    return d


def _verdict_fill_for_labels(competitor_labels: list[str]):
    """Build a verdict-cell color rule that recognizes named competitors.

    With N competitors, a verdict cell may contain a competitor's full_name
    (= Philips lost on that row). Default `xlsx_styles.verdict_fill` only
    knows the four legacy strings, so we wrap it.
    """
    competitor_lookup = {l.strip().lower() for l in competitor_labels if l}

    def fill(value: str):
        v = (value or "").strip().lower()
        if not v:
            return None
        base = xlsx_styles.verdict_fill(value)
        if base is not None:
            return base
        if v in competitor_lookup:
            return xlsx_styles.LOSE_FILL
        return None

    return fill


def _wb_to_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def comparison_to_excel(
    comparison: Comparison, label_a: str, competitor_labels: list[str]
) -> bytes:
    """Single-workbook styled output for flat (legacy / Custom View) comparisons.

    `competitor_labels` lists every competitor (one or more), in the same
    order their values appear in each ComparisonRow.competitor_values.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    base_keys = ["group", "feature", "philips_value", "competitor_value",
                 "verdict", "notes"]
    base_headers = ["Group", "Feature", "Philips", "Competitor",
                    "Verdict", "Notes"]
    expanded_keys, expanded_headers = expand_columns_for_competitors(
        base_keys, base_headers, competitor_labels,
    )

    sheet = wb.create_sheet("Comparison")
    xlsx_styles.write_styled_sheet(
        sheet,
        title="Comparison",
        subtitle=xlsx_styles.subtitle_for(label_a, competitor_labels),
        column_keys=expanded_keys,
        column_headers=_resolve_headers(expanded_headers, label_a, competitor_labels),
        rows=[_comparison_row_to_dict(r) for r in comparison.rows],
        color_rules={"verdict": _verdict_fill_for_labels(competitor_labels)},
    )

    summary_sheet = wb.create_sheet("Summary")
    xlsx_styles.write_narrative_sheet(
        summary_sheet,
        title="Summary",
        subtitle=xlsx_styles.subtitle_for(label_a, competitor_labels),
        narrative=comparison.summary,
    )
    return _wb_to_bytes(wb)


def block_to_xlsx(
    spec: BlockSpec,
    content: BlockContent,
    label_a: str,
    competitor_labels: list[str],
) -> bytes:
    """Render a single block as a standalone xlsx (marketing per-block downloads)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet = wb.create_sheet(_safe_sheet_name(spec.name))
    expanded_keys, expanded_headers = expand_columns_for_competitors(
        spec.columns, spec.headers, competitor_labels,
    )
    xlsx_styles.write_styled_sheet(
        sheet,
        title=spec.name,
        subtitle=xlsx_styles.subtitle_for(label_a, competitor_labels),
        column_keys=expanded_keys,
        column_headers=_resolve_headers(expanded_headers, label_a, competitor_labels),
        rows=content.rows,
        color_rules={
            "verdict": _verdict_fill_for_labels(competitor_labels),
            "winner": _verdict_fill_for_labels(competitor_labels),
            "who_tells_better": _verdict_fill_for_labels(competitor_labels),
        },
    )
    return _wb_to_bytes(wb)


def blocks_to_workbook(
    specs: list[BlockSpec],
    contents: list[BlockContent],
    summary_narrative: str,
    label_a: str,
    competitor_labels: list[str],
    summary_sheet_name: str = "Recommended for next launch",
) -> bytes:
    """Render every block + optional narrative-summary sheet into one workbook (R&D)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    color_rules = {
        "verdict": _verdict_fill_for_labels(competitor_labels),
        "winner": _verdict_fill_for_labels(competitor_labels),
        "who_tells_better": _verdict_fill_for_labels(competitor_labels),
    }
    contents_by_key = {c.block_key: c for c in contents}
    for spec in specs:
        content = contents_by_key.get(spec.key) or BlockContent(
            block_key=spec.key, narrative="", rows=[]
        )
        sheet = wb.create_sheet(_safe_sheet_name(spec.name))
        expanded_keys, expanded_headers = expand_columns_for_competitors(
            spec.columns, spec.headers, competitor_labels,
        )
        xlsx_styles.write_styled_sheet(
            sheet,
            title=spec.name,
            subtitle=xlsx_styles.subtitle_for(label_a, competitor_labels),
            column_keys=expanded_keys,
            column_headers=_resolve_headers(expanded_headers, label_a, competitor_labels),
            rows=content.rows,
            color_rules=color_rules,
        )

    if summary_narrative and summary_narrative.strip():
        sheet = wb.create_sheet(_safe_sheet_name(summary_sheet_name))
        xlsx_styles.write_narrative_sheet(
            sheet,
            title=summary_sheet_name,
            subtitle=xlsx_styles.subtitle_for(label_a, competitor_labels),
            narrative=summary_narrative,
        )
    return _wb_to_bytes(wb)


def generate_comparison(
    model_a: str, competitor_models: list[str], template: str
) -> bytes:
    """Compare one Philips/OBM monitor against N competitors and return .xlsx bytes.

    The returned workbook has a "Comparison" sheet (one row per feature
    with the winner named in the Verdict column) plus a "Summary" sheet.
    """
    comparison, label_a, competitor_labels = compute_comparison(
        model_a, competitor_models, template,
    )
    return comparison_to_excel(comparison, label_a, competitor_labels)
