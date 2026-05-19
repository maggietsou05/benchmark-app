"""
xlsx_styles.py — shared styling for benchmark Excel exports.

Every comparison workbook (R&D, B2C marketing, B2B marketing, Custom) calls
write_styled_sheet() so the look is identical across views. Colors follow
the MMD brand palette (matching style.py for app-vs-download consistency)
and the manager-approved verdict scheme from
example/benchmark_pipeline_handoff.md.
"""

from datetime import date

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ─── Brand palette (mirrors style.py for app/download consistency) ─

BRAND_NAVY = "0C5A8C"        # primary header fill — "navy" per user
BRAND_BLUE_TINT = "F5FAFE"   # row stripe
BRAND_BLUE_BORDER = "B5D4F4"

# Semantic fills for verdicts and Y/N cells — light backgrounds, dark text
WIN_FILL = "DBEFCB"          # Philips / Yes / Better — soft green
LOSE_FILL = "F8CFC9"         # Competitor / No / Lower — soft red
TIE_FILL = "D6E7F7"          # Tie — soft blue
INVESTIGATE_FILL = "FBE5BE"  # Investigate / Manual — soft amber

# Change-type badge fills (R&D blocks)
CHANGE_TYPE_FILLS = {
    "firmware": "E0E0E0",       # grey
    "cert": "E8D7F2",           # light purple
    "port swap": "C6E5DD",      # teal
    "new panel": "FCD9A8",      # orange
    "new platform": "F5B884",   # dark orange
}


# ─── Fonts, fills, borders, alignments ─────────────────────────────

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color=BRAND_NAVY)
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="595959")
BODY_FONT = Font(name="Calibri", size=10, color="333333")
NARRATIVE_FONT = Font(name="Calibri", size=11, color="333333")

HEADER_FILL = PatternFill("solid", fgColor=BRAND_NAVY)
STRIPE_FILL = PatternFill("solid", fgColor=BRAND_BLUE_TINT)

THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)


# ─── Column-value → fill colour rules ──────────────────────────────

def verdict_fill(value: str) -> str | None:
    v = (value or "").strip().lower()
    if v == "philips":
        return WIN_FILL
    if v == "competitor":
        return LOSE_FILL
    if v == "tie":
        return TIE_FILL
    if v == "investigate":
        return INVESTIGATE_FILL
    return None


def yes_no_fill(value: str) -> str | None:
    v = (value or "").strip().lower()
    if v in {"yes", "y", "true", "✓", "lists it", "listed"}:
        return WIN_FILL
    if v in {"no", "n", "false", "✗", "—", "-", "not listed", "silent"}:
        return LOSE_FILL
    return None


def change_type_fill(value: str) -> str | None:
    v = (value or "").strip().lower()
    for key, fill in CHANGE_TYPE_FILLS.items():
        if key in v:
            return fill
    return None


def direction_fill(value: str) -> str | None:
    """Indicate which direction is better for the spec (higher / lower / manual)."""
    v = (value or "").strip().lower()
    if "higher" in v:
        return WIN_FILL
    if "lower" in v:
        return LOSE_FILL
    if "manual" in v or "review" in v:
        return INVESTIGATE_FILL
    return None


# Column-key → cell coloring rule. Used by default; templates can override.
DEFAULT_COLOR_RULES = {
    "verdict": verdict_fill,
    "philips_lists_it": yes_no_fill,
    "competitor_lists_it": yes_no_fill,
    "philips_url_covers": yes_no_fill,
    "competitor_url_covers": yes_no_fill,
    "change_type": change_type_fill,
    "direction": direction_fill,
}


# ─── Column widths ─────────────────────────────────────────────────

COLUMN_WIDTHS = {
    "feature": 32,
    "spec": 32,
    "dimension": 22,
    "use_case": 28,
    "philips_value": 36,
    "competitor_value": 36,
    "philips_claim": 40,
    "competitor_claim": 40,
    "philips_url_covers": 18,
    "competitor_url_covers": 18,
    "philips_lists_it": 18,
    "competitor_lists_it": 18,
    "verdict": 14,
    "winner": 30,
    "direction": 22,
    "change_type": 18,
    "why_it_matters": 44,
    "who_tells_better": 22,
    "narrative_gap": 50,
    "notes": 50,
    "rationale": 50,
    "source_block": 22,
    "group": 22,
}
DEFAULT_COLUMN_WIDTH = 28


# ─── The writers ───────────────────────────────────────────────────

TITLE_ROW = 1
SUBTITLE_ROW = 2
HEADER_ROW = 4
BODY_START_ROW = 5


def write_styled_sheet(
    ws: Worksheet,
    title: str,
    subtitle: str,
    column_keys: list[str],
    column_headers: list[str],
    rows: list[dict],
    color_rules: dict | None = None,
) -> None:
    """Write a styled block sheet: title + subtitle + header + body rows.

    title        — block name displayed in the title row
    subtitle     — context line (e.g. "Philips X vs Competitor Y — 2026-05-11")
    column_keys  — ordered keys to look up in each row dict
    column_headers — parallel list of display headers
    rows         — list of dicts; missing keys render blank
    color_rules  — optional overrides merged onto DEFAULT_COLOR_RULES
    """
    if len(column_keys) != len(column_headers):
        raise ValueError("column_keys and column_headers must be the same length.")

    rules = {**DEFAULT_COLOR_RULES, **(color_rules or {})}
    n_cols = max(len(column_keys), 1)

    # Title and subtitle rows, each spanning every column for a clean look
    title_cell = ws.cell(row=TITLE_ROW, column=1, value=title)
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    if n_cols > 1:
        ws.merge_cells(
            start_row=TITLE_ROW, start_column=1,
            end_row=TITLE_ROW, end_column=n_cols,
        )
    ws.row_dimensions[TITLE_ROW].height = 26

    subtitle_cell = ws.cell(row=SUBTITLE_ROW, column=1, value=subtitle)
    subtitle_cell.font = SUBTITLE_FONT
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
    if n_cols > 1:
        ws.merge_cells(
            start_row=SUBTITLE_ROW, start_column=1,
            end_row=SUBTITLE_ROW, end_column=n_cols,
        )
    ws.row_dimensions[SUBTITLE_ROW].height = 18

    # Row 3 is intentionally blank (visual breathing room)

    # Header row
    for col_idx, header in enumerate(column_headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[HEADER_ROW].height = 30

    # Body rows
    for r_idx, row in enumerate(rows):
        excel_row = BODY_START_ROW + r_idx
        stripe = (r_idx % 2 == 1)
        for c_idx, col_key in enumerate(column_keys, start=1):
            raw = row.get(col_key, "") if isinstance(row, dict) else ""
            value = "" if raw is None else str(raw)
            cell = ws.cell(row=excel_row, column=c_idx, value=value)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = THIN_BORDER

            # Semantic colour (verdict / Y-N / change-type) wins over stripe
            colour = None
            rule = rules.get(col_key)
            if rule:
                colour = rule(value)
            if colour:
                cell.fill = PatternFill("solid", fgColor=colour)
            elif stripe:
                cell.fill = STRIPE_FILL

    # Freeze the header so it stays visible when scrolling
    ws.freeze_panes = ws.cell(row=BODY_START_ROW, column=1)

    # Column widths
    for col_idx, col_key in enumerate(column_keys, start=1):
        width = COLUMN_WIDTHS.get(col_key, DEFAULT_COLUMN_WIDTH)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_narrative_sheet(
    ws: Worksheet,
    title: str,
    subtitle: str,
    narrative: str,
) -> None:
    """Write a sheet containing only a title and narrative paragraph.

    Used for R&D Block 4 (Recommended for next launch) which is narrative-only.
    """
    title_cell = ws.cell(row=TITLE_ROW, column=1, value=title)
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[TITLE_ROW].height = 26

    subtitle_cell = ws.cell(row=SUBTITLE_ROW, column=1, value=subtitle)
    subtitle_cell.font = SUBTITLE_FONT
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[SUBTITLE_ROW].height = 18

    cell = ws.cell(row=HEADER_ROW, column=1, value=narrative or "")
    cell.font = NARRATIVE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    # Rough auto-height: 22 px per logical line
    lines = max(narrative.count("\n") + 1, 4) if narrative else 4
    ws.row_dimensions[HEADER_ROW].height = min(22 * lines + 20, 600)
    ws.column_dimensions["A"].width = 110


def subtitle_for(label_a: str, label_b: str) -> str:
    """Standard subtitle string for every sheet."""
    return f"{label_a} vs {label_b} — {date.today().isoformat()}"
